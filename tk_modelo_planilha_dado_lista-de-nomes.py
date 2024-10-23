import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import pandas as pd
import ttkbootstrap as ttk

# Função para abrir o arquivo
def selecionar_arquivo(entry):
    filepath = filedialog.askopenfilename()
    entry.delete(0, tk.END)
    entry.insert(0, filepath)

# Função para obter o formato do arquivo
def obter_formato():
    return var_formato.get()

# Função para verificar se uma célula é a principal de uma mesclagem
def is_top_left_of_merged_cell(sheet, cell):
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # Retorna True apenas se for a célula superior esquerda (início) do range
            if cell.coordinate == merged_range.coord.split(':')[0]:
                return True
    return False

# Função para copiar formatação e conteúdo de uma aba (XLSX)
def copiar_formato_e_conteudo(model_sheet, new_sheet):
    # Copiar dimensões das colunas
    for col in model_sheet.column_dimensions:
        new_sheet.column_dimensions[col].width = model_sheet.column_dimensions[col].width
    
    # Copiar altura das linhas
    for row in model_sheet.row_dimensions:
        new_sheet.row_dimensions[row].height = model_sheet.row_dimensions[row].height
    
    # Copiar células mescladas
    if model_sheet.merged_cells:
        for merged_range in model_sheet.merged_cells.ranges:
            new_sheet.merge_cells(str(merged_range))
    
    # Copiar estilos e valores de cada célula
    for row in model_sheet.iter_rows():
        for cell in row:
            new_cell = new_sheet[cell.coordinate]
            
            # Copiar o conteúdo apenas se for a célula superior esquerda de uma mesclagem
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                new_cell.value = cell.value  # Copiar conteúdo da célula original
            
            # Copiar estilos de célula
            if cell.has_style:
                new_cell.font = openpyxl.styles.Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    vertAlign=cell.font.vertAlign,
                    underline=cell.font.underline,
                    strike=cell.font.strike,
                    color=cell.font.color
                )
                new_cell.border = openpyxl.styles.Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                new_cell.fill = openpyxl.styles.PatternFill(
                    fill_type=cell.fill.fill_type,
                    start_color=cell.fill.start_color,
                    end_color=cell.fill.end_color
                )
                new_cell.alignment = openpyxl.styles.Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    text_rotation=cell.alignment.text_rotation,
                    wrap_text=cell.alignment.wrap_text,
                    shrink_to_fit=cell.alignment.shrink_to_fit,
                    indent=cell.alignment.indent
                )
                new_cell.number_format = cell.number_format

# Função para clonar a formatação da aba original nas novas abas
def clonar_formatacao_para_abas(workbook, model_sheet, nomes, area_responsavel):
    for i, sheet in enumerate(workbook.worksheets[1:], 1):
        copiar_formato_e_conteudo(model_sheet, sheet)

        # Definir B1 com o valor do arquivo area.txt
        sheet["B1"].value = area_responsavel
        
        # Definir B2 com o nome completo de lista_nomes.txt
        sheet["B2"].value = nomes[i - 1]  # Nome completo de cada aba
        
# Função para criar a planilha
def criar_planilha():
    formato = obter_formato()
    modelo_file = entry_modelo.get()
    nomes_file = entry_lista_nomes.get()
    area_file = entry_area.get()

    if not (modelo_file and nomes_file and area_file):
        messagebox.showerror("Erro", "Por favor, selecione todos os arquivos necessários.")
        return

    # Ler a lista de nomes usando pandas
    nomes = pd.read_csv(nomes_file, header=None)[0].tolist()

    # Ler o conteúdo do arquivo area.txt
    with open(area_file, 'r', encoding='utf-8') as f:
        area_responsavel = f.read().strip()

    if formato == 'xlsx':
        # Carregar o modelo de planilha (XLSX)
        workbook = openpyxl.load_workbook(modelo_file)
        model_sheet = workbook.active  # Considerando que a primeira aba seja o modelo

        # Criar abas com os primeiros nomes
        for nome_completo in nomes:
            workbook.create_sheet(title=nome_completo.split()[0])  # Criar uma aba para cada nome

        # Clonar a formatação e inserir dados nas abas novas
        clonar_formatacao_para_abas(workbook, model_sheet, nomes, area_responsavel)

        # Salvar a planilha
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("XLSX files", "*.xlsx")])

        if save_path:
            workbook.save(save_path)
            messagebox.showinfo("Sucesso", "Planilha criada com sucesso!")

# Interface gráfica com ttkbootstrap
root = ttk.Window(themename="cosmo")
root.title("Gerador de Planilha")
root.geometry("600x400")
root.resizable(True, True)

# Formato do arquivo
var_formato = tk.StringVar(value="xlsx")
ttk.Label(root, text="Escolha o formato de arquivo:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
ttk.Radiobutton(root, text="XLSX", variable=var_formato, value="xlsx").grid(row=1, column=0, padx=10, pady=5, sticky="w")

# Seleção de arquivos
ttk.Label(root, text="Arquivo modelo de planilha:").grid(row=2, column=0, padx=10, pady=5, sticky="w")
entry_modelo = ttk.Entry(root, width=40)
entry_modelo.grid(row=3, column=0, padx=10, pady=5, sticky="we")
btn_modelo = ttk.Button(root, text="Selecionar", command=lambda: selecionar_arquivo(entry_modelo))
btn_modelo.grid(row=3, column=1, padx=10, pady=5)

ttk.Label(root, text="Arquivo lista_nomes.txt:").grid(row=4, column=0, padx=10, pady=5, sticky="w")
entry_lista_nomes = ttk.Entry(root, width=40)
entry_lista_nomes.grid(row=5, column=0, padx=10, pady=5, sticky="we")
btn_lista_nomes = ttk.Button(root, text="Selecionar", command=lambda: selecionar_arquivo(entry_lista_nomes))
btn_lista_nomes.grid(row=5, column=1, padx=10, pady=5)

ttk.Label(root, text="Arquivo area.txt:").grid(row=6, column=0, padx=10, pady=5, sticky="w")
entry_area = ttk.Entry(root, width=40)
entry_area.grid(row=7, column=0, padx=10, pady=5, sticky="we")
btn_area = ttk.Button(root, text="Selecionar", command=lambda: selecionar_arquivo(entry_area))
btn_area.grid(row=7, column=1, padx=10, pady=5)

# Botão para criar a planilha
btn_criar = ttk.Button(root, text="Criar Planilha", command=criar_planilha)
btn_criar.grid(row=8, column=0, columnspan=2, padx=10, pady=10)

# Configurar redimensionamento
root.columnconfigure(0, weight=1)

root.mainloop()
